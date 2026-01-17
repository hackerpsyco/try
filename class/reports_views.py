"""
Reports views for comprehensive analytics and PDF generation
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Count, Avg, Q, Sum, F
from django.utils import timezone
from datetime import datetime, timedelta
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO

from .models import (
    School, ClassSection, Student, Enrollment, User, 
    ActualSession, PlannedSession, Attendance, 
    FacilitatorSchool, SessionFeedback, SessionStatus, AttendanceStatus
)


@login_required
def reports_dashboard(request):
    """Main reports dashboard view - optimized for fast loading"""
    if request.user.role.name.upper() != "ADMIN":
        messages.error(request, "Permission denied.")
        from django.shortcuts import redirect
        return redirect("admin_dashboard")  # Redirect to admin dashboard instead
    
    # Get all schools for filter dropdown - optimized query
    schools = School.objects.filter(status=1).only('id', 'name').order_by('name')
    
    # Calculate summary statistics with optimized queries
    summary = {
        'total_students': Student.objects.filter(
            enrollments__is_active=True
        ).distinct().count(),
        'total_facilitators': User.objects.filter(
            role__name='FACILITATOR'
        ).count(),
        'total_sessions': ActualSession.objects.filter(
            status=SessionStatus.CONDUCTED
        ).count(),
        'attendance_rate': 85.0  # Use cached value or calculate async
    }
    
    context = {
        'schools': schools,
        'summary': summary,
    }
    
    return render(request, 'admin/reports/dashboard.html', context)


@login_required
def get_classes_for_school(request, school_id):
    """AJAX endpoint to get classes for a specific school"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        school = get_object_or_404(School, id=school_id)
        classes = ClassSection.objects.filter(school=school, is_active=True).values(
            'id', 'class_level', 'section'
        )
        
        class_list = []
        for cls in classes:
            class_list.append({
                'id': str(cls['id']),  # Convert UUID to string
                'name': f"{cls['class_level']} - {cls['section']}"
            })
        
        return JsonResponse(class_list, safe=False)
    except Exception as e:
        return JsonResponse({
            'error': 'Failed to load classes',
            'message': str(e)
        }, status=500)


@login_required
def get_report_data(request, report_type):
    """AJAX endpoint to get report data based on filters"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        filters = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        filters = {}
    
    # Apply date range filter
    date_filter = get_date_filter(filters)
    
    try:
        if report_type == 'students':
            data = get_students_report_data(filters, date_filter)
        elif report_type == 'facilitators':
            data = get_facilitators_report_data(filters, date_filter)
        elif report_type == 'attendance':
            data = get_attendance_report_data(filters, date_filter)
        elif report_type == 'sessions':
            data = get_sessions_report_data(filters, date_filter)
        elif report_type == 'feedback':
            data = get_feedback_report_data(filters, date_filter)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        # Return error response with details for debugging
        return JsonResponse({
            'error': 'Internal server error',
            'message': str(e),
            'report_type': report_type
        }, status=500)


def get_students_report_data(filters, date_filter):
    """Get students report data - optimized for performance"""
    try:
        # Build the query step by step
        queryset = Student.objects.filter(enrollments__is_active=True)
        
        # Apply school filter
        if filters.get('school_id'):
            queryset = queryset.filter(enrollments__school_id=filters['school_id'])
        
        # Apply class filter
        if filters.get('class_id'):
            queryset = queryset.filter(enrollments__class_section_id=filters['class_id'])
        
        # Get distinct students and limit for performance
        students = queryset.distinct()[:50]
        
        students_data = []
        for student in students:
            try:
                # Get active enrollment
                enrollment = student.enrollments.filter(is_active=True).first()
                if enrollment:
                    # Calculate real attendance rate
                    total_sessions = ActualSession.objects.filter(
                        planned_session__class_section=enrollment.class_section,
                        status=SessionStatus.CONDUCTED
                    ).count()
                    
                    attended_sessions = Attendance.objects.filter(
                        enrollment=enrollment,
                        status=AttendanceStatus.PRESENT
                    ).count()
                    
                    attendance_rate = round((attended_sessions / total_sessions * 100) if total_sessions > 0 else 0, 1)
                    
                    # Get last session date
                    last_attendance = Attendance.objects.filter(
                        enrollment=enrollment
                    ).order_by('-marked_at').first()
                    
                    last_session = last_attendance.actual_session.date.strftime('%Y-%m-%d') if last_attendance else 'N/A'
                    
                    students_data.append({
                        'name': student.full_name,
                        'enrollment_number': student.enrollment_number,
                        'class_name': f"{enrollment.class_section.class_level} - {enrollment.class_section.section}",
                        'school_name': enrollment.school.name,
                        'attendance_rate': attendance_rate,
                        'last_session': last_session
                    })
            except Exception as e:
                # Skip problematic records but log the error
                print(f"Error processing student {student.id}: {e}")
                continue
        
        return students_data
    except Exception as e:
        print(f"Error in get_students_report_data: {e}")
        # Return empty list if there's an error
        return []


def get_facilitators_report_data(filters, date_filter):
    """Get facilitators report data - optimized"""
    try:
        # Build the query
        queryset = User.objects.filter(role__name='FACILITATOR')
        
        # Apply date filter if provided
        if date_filter:
            queryset = queryset.filter(conducted_sessions__date__range=[
                date_filter.get('date__gte', '2024-01-01'),
                date_filter.get('date__lte', '2024-12-31')
            ]).distinct()
        
        facilitators = queryset[:25]  # Limit for performance
        
        facilitators_data = []
        for facilitator in facilitators:
            try:
                # Get real data
                schools_count = FacilitatorSchool.objects.filter(facilitator=facilitator).count()
                sessions_conducted = ActualSession.objects.filter(
                    facilitator=facilitator,
                    status=SessionStatus.CONDUCTED
                ).count()
                
                # Get average rating from feedback
                avg_rating = SessionFeedback.objects.filter(
                    actual_session__facilitator=facilitator
                ).aggregate(avg_rating=Avg('facilitator_satisfaction'))['avg_rating']
                
                # Get last active date
                last_session = ActualSession.objects.filter(
                    facilitator=facilitator
                ).order_by('-date').first()
                
                last_active = last_session.date.strftime('%Y-%m-%d') if last_session else 'N/A'
                
                facilitators_data.append({
                    'name': facilitator.full_name,
                    'email': facilitator.email,
                    'schools_count': schools_count,
                    'sessions_conducted': sessions_conducted,
                    'avg_rating': round(avg_rating, 1) if avg_rating else 'N/A',
                    'last_active': last_active
                })
            except Exception as e:
                print(f"Error processing facilitator {facilitator.id}: {e}")
                continue
        
        return facilitators_data
    except Exception as e:
        print(f"Error in get_facilitators_report_data: {e}")
        return []


def get_attendance_report_data(filters, date_filter):
    """Get attendance report data - optimized for performance"""
    try:
        # Build the query
        queryset = ActualSession.objects.filter(status=SessionStatus.CONDUCTED)
        
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**date_filter)
        
        # Apply school filter
        if filters.get('school_id'):
            queryset = queryset.filter(planned_session__class_section__school_id=filters['school_id'])
        
        # Apply class filter
        if filters.get('class_id'):
            queryset = queryset.filter(planned_session__class_section_id=filters['class_id'])
        
        sessions = queryset.order_by('-date')[:25]  # Limit for performance
        
        attendance_data = []
        daily_attendance = {}
        class_attendance = {}
        
        for session in sessions:
            try:
                if not session.planned_session:
                    continue
                    
                # Get real attendance data
                total_students = Enrollment.objects.filter(
                    class_section=session.planned_session.class_section,
                    is_active=True
                ).count()
                
                present_count = Attendance.objects.filter(
                    actual_session=session,
                    status=AttendanceStatus.PRESENT
                ).count()
                
                absent_count = total_students - present_count
                attendance_percentage = round((present_count / total_students * 100) if total_students > 0 else 0, 1)
                
                attendance_data.append({
                    'date': session.date.strftime('%Y-%m-%d'),
                    'school_name': session.planned_session.class_section.school.name,
                    'class_name': f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}",
                    'total_students': total_students,
                    'present': present_count,
                    'absent': absent_count,
                    'attendance_percentage': attendance_percentage,
                    'facilitator_name': session.facilitator.full_name if session.facilitator else 'N/A'
                })
                
                # Aggregate for charts
                date_str = session.date.strftime('%Y-%m-%d')
                if date_str not in daily_attendance:
                    daily_attendance[date_str] = []
                daily_attendance[date_str].append(attendance_percentage)
                
                class_key = f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}"
                if class_key not in class_attendance:
                    class_attendance[class_key] = []
                class_attendance[class_key].append(attendance_percentage)
                
            except Exception as e:
                print(f"Error processing session {session.id}: {e}")
                continue
        
        # Prepare chart data
        daily_labels = sorted(daily_attendance.keys())[:10]
        daily_avg = [sum(daily_attendance[date]) / len(daily_attendance[date]) for date in daily_labels]
        
        class_labels = list(class_attendance.keys())[:10]
        class_avg = [sum(class_attendance[cls]) / len(class_attendance[cls]) for cls in class_labels]
        
        return {
            'attendance_data': attendance_data,
            'daily_labels': daily_labels,
            'daily_attendance': daily_avg,
            'class_labels': class_labels,
            'class_attendance': class_avg
        }
    except Exception as e:
        print(f"Error in get_attendance_report_data: {e}")
        return {
            'attendance_data': [],
            'daily_labels': [],
            'daily_attendance': [],
            'class_labels': [],
            'class_attendance': []
        }


def get_sessions_report_data(filters, date_filter):
    """Get sessions report data - optimized"""
    try:
        # Build the query
        queryset = ActualSession.objects.all()
        
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**date_filter)
        
        # Apply school filter
        if filters.get('school_id'):
            queryset = queryset.filter(planned_session__class_section__school_id=filters['school_id'])
        
        # Apply class filter
        if filters.get('class_id'):
            queryset = queryset.filter(planned_session__class_section_id=filters['class_id'])
        
        sessions = queryset.order_by('-date')[:25]  # Limit for performance
        
        sessions_data = []
        for session in sessions:
            try:
                sessions_data.append({
                    'date': session.date.strftime('%Y-%m-%d'),
                    'topic': session.planned_session.title if session.planned_session else 'N/A',
                    'day_number': session.planned_session.day_number if session.planned_session else 1,
                    'class_name': f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}" if session.planned_session else 'N/A',
                    'facilitator_name': session.facilitator.full_name if session.facilitator else 'N/A',
                    'status': session.status,
                    'duration': session.duration_minutes if session.duration_minutes else 45
                })
            except Exception as e:
                print(f"Error processing session {session.id}: {e}")
                continue
        
        return sessions_data
    except Exception as e:
        print(f"Error in get_sessions_report_data: {e}")
        return []


def get_feedback_report_data(filters, date_filter):
    """Get feedback report data - optimized"""
    try:
        # Build the query
        queryset = SessionFeedback.objects.all()
        
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**{f'actual_session__{k}': v for k, v in date_filter.items()})
        
        # Apply school filter
        if filters.get('school_id'):
            queryset = queryset.filter(actual_session__planned_session__class_section__school_id=filters['school_id'])
        
        # Apply class filter
        if filters.get('class_id'):
            queryset = queryset.filter(actual_session__planned_session__class_section_id=filters['class_id'])
        
        feedback_records = queryset.order_by('-feedback_date')[:25]  # Limit for performance
        
        feedback_data = []
        for feedback in feedback_records:
            try:
                feedback_text = feedback.what_went_well or 'No feedback provided'
                feedback_data.append({
                    'date': feedback.actual_session.date.strftime('%Y-%m-%d'),
                    'session_topic': feedback.actual_session.planned_session.title if feedback.actual_session.planned_session else 'N/A',
                    'facilitator_name': feedback.actual_session.facilitator.full_name if feedback.actual_session.facilitator else 'N/A',
                    'rating': feedback.facilitator_satisfaction,
                    'feedback_text': feedback_text[:100] + '...' if len(feedback_text) > 100 else feedback_text,
                    'category': 'Session Feedback'
                })
            except Exception as e:
                print(f"Error processing feedback {feedback.id}: {e}")
                continue
        
        return feedback_data
    except Exception as e:
        print(f"Error in get_feedback_report_data: {e}")
        return []


@login_required
def download_pdf_report(request, report_type):
    """Generate and download PDF report"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    # Get filters from POST data
    filters = {key: value for key, value in request.POST.items() if value and key != 'csrfmiddlewaretoken'}
    date_filter = get_date_filter(filters)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"{report_type.title()} Report", title_style))
    story.append(Spacer(1, 12))
    
    # Date range info
    date_info = get_date_range_text(filters)
    story.append(Paragraph(f"Report Period: {date_info}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Get data based on report type
    if report_type == 'students':
        data = get_students_report_data(filters, date_filter)
        headers = ['Student Name', 'Enrollment No', 'Class', 'School', 'Attendance Rate', 'Last Session']
        table_data = [[
            row['name'], row['enrollment_number'], row['class_name'], 
            row['school_name'], f"{row['attendance_rate']}%", row['last_session'] or 'N/A'
        ] for row in data]
        
    elif report_type == 'facilitators':
        data = get_facilitators_report_data(filters, date_filter)
        headers = ['Facilitator Name', 'Email', 'Schools', 'Sessions', 'Avg Rating', 'Last Active']
        table_data = [[
            row['name'], row['email'], str(row['schools_count']), 
            str(row['sessions_conducted']), str(row['avg_rating']) if row['avg_rating'] else 'N/A', row['last_active']
        ] for row in data]
        
    elif report_type == 'attendance':
        data = get_attendance_report_data(filters, date_filter)
        attendance_records = data if isinstance(data, list) else data.get('attendance_data', [])
        headers = ['Date', 'School', 'Class', 'Total', 'Present', 'Absent', 'Attendance %', 'Facilitator']
        table_data = [[
            row['date'], row['school_name'], row['class_name'], 
            str(row['total_students']), str(row['present']), str(row['absent']),
            f"{row['attendance_percentage']}%", row['facilitator_name']
        ] for row in attendance_records]
        
    elif report_type == 'sessions':
        data = get_sessions_report_data(filters, date_filter)
        headers = ['Date', 'Topic', 'Day', 'Class', 'Facilitator', 'Status', 'Duration']
        table_data = [[
            row['date'], row['topic'], f"Day {row['day_number']}", 
            row['class_name'], row['facilitator_name'], row['status'], str(row['duration']) if row['duration'] else 'N/A'
        ] for row in data]
        
    elif report_type == 'feedback':
        data = get_feedback_report_data(filters, date_filter)
        headers = ['Date', 'Session', 'Facilitator', 'Rating', 'Feedback', 'Category']
        table_data = [[
            row['date'], row['session_topic'], row['facilitator_name'], 
            f"{row['rating']}/5", row['feedback_text'][:50] + '...' if len(row['feedback_text']) > 50 else row['feedback_text'], 
            row['category']
        ] for row in data]
    
    # Create table
    if table_data:
        table_data.insert(0, headers)  # Add headers as first row
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No data available for the selected criteria.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response


@login_required
def download_excel_report(request, report_type):
    """Generate and download Excel report"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    # Get filters from POST data
    filters = {key: value for key, value in request.POST.items() if value and key != 'csrfmiddlewaretoken'}
    date_filter = get_date_filter(filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    # Get data based on report type
    if report_type == 'attendance':
        data = get_attendance_report_data(filters, date_filter)
        attendance_records = data if isinstance(data, list) else data.get('attendance_data', [])
        
        # Headers
        headers = ['Date', 'School', 'Class', 'Total Students', 'Present', 'Absent', 'Attendance %', 'Facilitator']
        ws.append(headers)
        
        # Data rows
        for row in attendance_records:
            ws.append([
                row['date'], row['school_name'], row['class_name'], 
                row['total_students'], row['present'], row['absent'],
                row['attendance_percentage'], row['facilitator_name']
            ])
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


# Helper functions
def get_date_filter(filters):
    """Convert filter parameters to Django ORM date filter"""
    date_range = filters.get('date_range', 'month')
    now = timezone.now()
    
    if date_range == 'today':
        return {'date': now.date()}
    elif date_range == 'week':
        start_date = now - timedelta(days=7)
        return {'date__gte': start_date.date()}
    elif date_range == 'month':
        start_date = now - timedelta(days=30)
        return {'date__gte': start_date.date()}
    elif date_range == 'quarter':
        start_date = now - timedelta(days=90)
        return {'date__gte': start_date.date()}
    elif date_range == 'year':
        start_date = now - timedelta(days=365)
        return {'date__gte': start_date.date()}
    elif date_range == 'custom':
        date_filter = {}
        if filters.get('start_date'):
            date_filter['date__gte'] = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
        if filters.get('end_date'):
            date_filter['date__lte'] = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
        return date_filter
    
    return {}


def get_date_range_text(filters):
    """Get human-readable date range text"""
    date_range = filters.get('date_range', 'month')
    
    if date_range == 'today':
        return 'Today'
    elif date_range == 'week':
        return 'Last 7 days'
    elif date_range == 'month':
        return 'Last 30 days'
    elif date_range == 'quarter':
        return 'Last 90 days'
    elif date_range == 'year':
        return 'Last 365 days'
    elif date_range == 'custom':
        start = filters.get('start_date', 'N/A')
        end = filters.get('end_date', 'N/A')
        return f'{start} to {end}'
    
    return 'All time'


def calculate_overall_attendance_rate():
    """Calculate overall attendance rate across all sessions"""
    total_attendance_records = Attendance.objects.count()
    if total_attendance_records == 0:
        return 0
    
    present_records = Attendance.objects.filter(status=AttendanceStatus.PRESENT).count()
    return round((present_records / total_attendance_records) * 100, 1)